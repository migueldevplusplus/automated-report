from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_formatted_excel_report(filepath: Path, tables_list: list):
    """
    Adapted version that receives an ordered list of (table_name, dataframe)
    and generates exactly the same layout as the original script.
    """
    SHEET = "dashboard_data"
    fixed_start_rows = [1, 31, 61, 91, 121, 151, 181, 211]

    # ── 1. Write data to fixed positions ─────────────────────────────────────────
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="w") as writer:
        for (table_name, df_table), start_row in zip(tables_list, fixed_start_rows):
            df_table.to_excel(writer, sheet_name=SHEET, startrow=start_row - 1, index=False)

    # ── 2. Load workbook and apply formatting ────────────────────────────────────
    wb = load_workbook(filepath)
    ws = wb[SHEET]

    # Column widths (exact match to original)
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20

    # ── 3. Create Excel tables with the same names ───────────────────────────────
    for (table_name, df_table), start_row in zip(tables_list, fixed_start_rows):
        if df_table.empty:
            continue

        num_rows = len(df_table)
        num_cols = df_table.shape[1]
        start_cell = f"A{start_row}"
        end_cell = f"{chr(64 + num_cols)}{start_row + num_rows - 1}"
        ref = f"{start_cell}:{end_cell}"

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )
        ws.add_table(table)

    # ── 4. Number formatting – EXACT COPY of original script ─────────────────────
    for table in ws.tables.values():
        start_cell, end_cell = table.ref.split(":")
        min_row = ws[start_cell].row + 1  # skip header
        max_row = ws[end_cell].row
        min_col = ws[start_cell].column
        max_col = ws[end_cell].column

        # Get headers
        header_cells = next(ws.iter_rows(
            min_row=min_row-1,
            max_row=min_row-1,
            min_col=min_col,
            max_col=max_col
        ))
        headers = [cell.value for cell in header_cells]

        value_col = headers.index("Value") + min_col if "Value" in headers else None
        percentage_col = headers.index("Percentage") + min_col if "Percentage" in headers else None

        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col
        ):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    if table.displayName == "tbl_top_performers" and cell.column == value_col:
                        row_offset = cell.row - min_row + 1
                        if row_offset <= 2:
                            cell.number_format = "#,##0.0"
                        else:
                            cell.number_format = "0.0%"
                    elif table.displayName in {"tbl_kpis_percentage_changes"} or \
                         (table.displayName == "tbl_payment_distribution" and cell.column == percentage_col):
                        cell.number_format = "0.0%"
                    elif table.displayName == "tbl_kpis" and cell.column == value_col:
                        metric = ws.cell(row=cell.row, column=min_col).value
                        if metric in {"Total Sales", "Average Ticket", "Gross Income"}:
                            cell.number_format = "#,##0.0"
                        elif metric in {"Transactions", "Total Quantity"}:
                            cell.number_format = "#,##0"
                        elif metric == "Average Rating":
                            cell.number_format = "0.0"
                    else:
                        cell.number_format = "#,##0.0" if not isinstance(cell.value, int) else "#,##0"

    wb.save(filepath)
    print(f"Excel report generated with identical layout to the original: {filepath}")